import openpyxl
from openpyxl.styles import Alignment, Border, Side


def export(kb_list, filename=""):
    """
    根据课表列表生成 Excel 文件（12节课 × 7天）。
    课表数据需要包含：jxbmc（课程名）、cdmc（教室）、jcor（节次）、xqj（星期几）
    """

    # 初始化 12x7 的课表二维数组（12 节课，7 天）
    schedule = [['' for _ in range(7)] for _ in range(12)]

    # 星期映射： "1" -> 周一, "2" -> 周二, ... "7" -> 周日
    weekday_mapping = {str(i): i - 1 for i in range(1, 8)}

    # 填充课表数据
    for course in kb_list:
        course_name = course.get('jxbmc', '')
        classroom = course.get('cdmc', '')
        jcor = course.get('jcor', '')
        weekday = weekday_mapping.get(course.get('xqj', ''), None)
        if not jcor or weekday is None:
            continue

        # 处理节次（可能为 "1-2" 或 "3"）
        periods = jcor.split('-')
        try:
            start_period = int(periods[0]) - 1
            end_period = int(periods[1]) - 1 if len(periods) > 1 else start_period
        except ValueError:
            continue

        for period in range(start_period, min(end_period + 1, 12)):
            schedule[period][weekday] = f"{course_name} ({classroom})"

    # 创建 Excel 工作簿和工作表
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "课程安排"

    # 添加表头
    headers = ["节次", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    ws.append(headers)

    # 添加课表数据行
    for i, row in enumerate(schedule):
        if i == 7:  # 第 8 节（索引为 7）
            ws.append([f"{i + 1} (6点下课)"] + row)
        else:
            ws.append([str(i + 1)] + row)

    # 设置自动调整列宽及启用文本换行
    wrap_alignment = Alignment(wrap_text=True)
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
            cell.alignment = wrap_alignment
        ws.column_dimensions[column_letter].width = max_length + 2

    # 调整行高
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 30

    # 添加粗边框
    thick_black = Side(border_style="thick", color="000000")

    # 在节次4和5之间、8和9之间加边框
    for col in range(1, 9):  # 从第1列到第8列（节次和周一到周日）
        ws.cell(row=5, column=col).border = Border(bottom=thick_black)  # 第5行下边框
        ws.cell(row=9, column=col).border = Border(bottom=thick_black)  # 第9行下边框

    # 保存文件
    wb.save(filename)
    print(f"Excel 文件已生成：{filename}")
